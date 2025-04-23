import serial
import threading
import time
from datetime import datetime
import os
import pandas as pd
import openpyxl
from openpyxl.comments import Comment
import tkinter as tk
from tkinter import filedialog
import serial.tools.list_ports
import serial

def obtener_puertos_usb(vid, pid):
    puertos = serial.tools.list_ports.comports()
    for puerto in puertos:
        # Extraer VID y PID desde el puerto en formato hexadecimal
        puerto_vid = puerto.vid
        puerto_pid = puerto.pid
        
        # Comparar VID y PID
        if puerto_vid == int(vid, 16) and puerto_pid == int(pid, 16):
            print(f"Puerto: {puerto.device}, VID: {puerto_vid}, PID: {puerto_pid}")
            return puerto.device  # Devuelve el puerto correspondiente
    return None

# Especificar el VID y PID de tu dispositivo
VID = "2341"
PID = "0042"

# Llamar a la función para obtener el puerto USB
PORT = obtener_puertos_usb(VID, PID)
if PORT is None:
    print("No se encontró el puerto del Arduino.")
else:
    print(f"Puerto USB encontrado: {PORT}")



BAUD_RATE = 9600
TIMEOUT = 1  # Timeout for serial reading in seconds

# Valores de longitud de onda proporcionados
WAVELENGTHS = [
    311.93, 314.64, 317.34, 320.05, 322.75, 325.45, 328.14, 330.84, 333.53, 336.21,
    338.90, 341.58, 344.26, 346.94, 349.61, 352.28, 354.95, 357.62, 360.28, 362.94,
    365.59, 368.24, 370.89, 373.54, 376.18, 378.82, 381.45, 384.08, 386.71, 389.34,
    391.96, 394.57, 397.19, 399.80, 402.40, 405.00, 407.60, 410.20, 412.79, 415.37,
    417.96, 420.53, 423.11, 425.68, 428.25, 430.81, 433.36, 435.92, 438.47, 441.01,
    443.55, 446.09, 448.62, 451.15, 453.67, 456.19, 458.70, 461.21, 463.71, 466.21,
    468.71, 471.20, 473.68, 476.16, 478.64, 481.11, 483.57, 486.04, 488.49, 490.94,
    493.39, 495.83, 498.27, 500.70, 503.12, 505.54, 507.96, 510.37, 512.77, 515.17,
    517.57, 519.96, 522.34, 524.72, 527.09, 529.46, 531.82, 534.18, 536.53, 538.88,
    541.22, 543.55, 545.88, 548.20, 550.52, 552.83, 555.14, 557.44, 559.74, 562.03,
    564.31, 566.59, 568.86, 571.13, 573.39, 575.64, 577.89, 580.13, 582.37, 584.60,
    586.83, 589.05, 591.26, 593.47, 595.67, 597.87, 600.06, 602.24, 604.42, 606.59,
    608.75, 610.91, 613.06, 615.21, 617.35, 619.49, 621.62, 623.74, 625.85, 627.96,
    630.07, 632.16, 634.26, 636.34, 638.42, 640.49, 642.56, 644.62, 646.67, 648.72,
    650.76, 652.79, 654.82, 656.84, 658.86, 660.87, 662.87, 664.87, 666.86, 668.84,
    670.82, 672.79, 674.76, 676.71, 678.67, 680.61, 682.55, 684.48, 686.41, 688.33,
    690.24, 692.15, 694.05, 695.94, 697.83, 699.71, 701.59, 703.46, 705.32, 707.18,
    709.02, 710.87, 712.70, 714.53, 716.36, 718.18, 719.99, 721.79, 723.59, 725.38,
    727.17, 728.94, 730.72, 732.48, 734.24, 736.00, 737.74, 739.48, 741.22, 742.95,
    744.67, 746.38, 748.09, 749.79, 751.49, 753.18, 754.86, 756.54, 758.21, 759.88,
    761.54, 763.19, 764.83, 766.47, 768.11, 769.73, 771.36, 772.97, 774.58, 776.18,
    777.78, 779.37, 780.95, 782.53, 784.11, 785.67, 787.23, 788.79, 790.33, 791.88,
    793.41, 794.94, 796.47, 797.98, 799.50, 801.00, 802.50, 804.00, 805.49, 806.97,
    808.45, 809.92, 811.39, 812.85, 814.30, 815.75, 817.19, 818.63, 820.06, 821.49,
    822.91, 824.32, 825.73, 827.13, 828.53, 829.92, 831.31, 832.69, 834.07, 835.44,
    836.81, 838.17, 839.52, 840.87, 842.22, 843.55, 844.89, 846.22, 847.54, 848.86,
    850.17, 851.48, 852.79, 854.08, 855.38, 856.67, 857.95, 859.23, 860.50, 861.77,
    863.04, 864.30, 865.55, 866.80, 868.05, 869.29, 870.53, 871.76, 872.99, 874.21,
    875.43, 876.64, 877.85, 879.06, 880.26, 881.46, 882.65, 883.84
]

class SerialMonitor:
    def __init__(self, port, baud_rate):
        self.port = port
        self.baud_rate = baud_rate
        self.serial_port = None
        self.running = False
        self.monitor_thread = None
        self.buffer = ""
        self.save_dir = None
        self.is_first_measurement = True
        self.measurement_count = 0
        self.current_batch = []
        self.manual_batch = []
        self.waiting_for_manual = False
        self.manual_name = None

    def set_save_directory(self):
        # Crear una ventana de Tkinter para abrir el explorador de archivos
        root = tk.Tk()
        root.withdraw()  # Ocultar la ventana principal
        self.save_dir = filedialog.askdirectory(title="Selecciona la carpeta para guardar los datos")
        if self.save_dir:
            print(f"Directorio seleccionado: {self.save_dir}")
        else:
            print("No se seleccionó ningún directorio. El programa se cerrará.")
            exit(1)

    def start(self):
        try:
            self.serial_port = serial.Serial(self.port, self.baud_rate, timeout=TIMEOUT)
            print(f"Puerto {self.port} abierto.")
            self.running = True
            self.monitor_thread = threading.Thread(target=self.monitor_serial, daemon=True)
            self.monitor_thread.start()
        except serial.SerialException as e:
            print(f"Error al abrir el puerto {self.port}: {e}")
            exit(1)

    def stop(self):
        self.running = False
        if self.monitor_thread:
            self.monitor_thread.join()
        if self.serial_port and self.serial_port.is_open:
            self.serial_port.close()
            print("Puerto cerrado.")

    def monitor_serial(self):
        while self.running:
            if self.serial_port.in_waiting > 0:
                data = self.serial_port.read(self.serial_port.in_waiting).decode('utf-8', errors='ignore')
                print(data, end='')
                self.buffer += data
                self.process_buffer()
            time.sleep(0.01)

    def process_buffer(self):
        while True:
            start_idx = self.buffer.find('h')
            end_idx = self.buffer.find('a', start_idx)
            if start_idx != -1 and end_idx != -1 and end_idx > start_idx:
                message = self.buffer[start_idx:end_idx + 1]
                self.buffer = self.buffer[end_idx + 1:]
                self.handle_arduino_message(message)
            else:
                break

    def handle_arduino_message(self, message):
        datos = message.split(',')
        
        if len(datos) < 4 or datos[0] != 'h' or datos[-1] != 'a':
            print("Formato de datos inválido.")
            return

        try:
            espectrometro = [float(x) for x in datos[1:-3] if x.strip()]
            temperatura = float(datos[-3])
            ph = float(datos[-2])
        except ValueError as e:
            print(f"Error al convertir datos a números: {e}")
            return

        if self.waiting_for_manual:
            self.manual_batch.append((espectrometro, temperatura, ph))
            print(f"Medición manual {len(self.manual_batch)}/10 recibida.")
            if len(self.manual_batch) == 10:
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                self.save_manual_batch(timestamp)
                self.manual_batch = []
                self.waiting_for_manual = False
                self.manual_name = None
        else:
            self.measurement_count += 1
            self.current_batch.append((espectrometro, temperatura, ph))
            print(f"Medición automática {self.measurement_count % 10 if self.measurement_count % 10 != 0 else 10}/10 recibida.")
            if self.measurement_count % 10 == 0:
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                tipo = "CALIBRACION DEL BLANCO" if self.is_first_measurement else "MEDICION"
                self.save_batch(timestamp, tipo)
                self.current_batch = []
                if self.is_first_measurement:
                    self.is_first_measurement = False
                    print("Calibración del blanco completada.")

    def save_batch(self, timestamp, tipo):
        # Crear DataFrame para el espectrómetro
        espectro_data = []
        for espectrometro, _, _ in self.current_batch:
            espectro_data.append(espectrometro)
        espectro_df = pd.DataFrame(espectro_data)

        # Crear DataFrame para la temperatura
        temp_data = [temp for _, temp, _ in self.current_batch]
        temp_df = pd.DataFrame(temp_data, columns=["Temperatura"])

        # Crear DataFrame para el pH
        ph_data = [ph for _, _, ph in self.current_batch]
        ph_df = pd.DataFrame(ph_data, columns=["pH"])

        # Guardar en archivos Excel
        espectro_file = os.path.join(self.save_dir, "valores_espectro.xlsx")
        temp_file = os.path.join(self.save_dir, "valores_temperatura.xlsx")
        ph_file = os.path.join(self.save_dir, "valores_ph.xlsx")

        # Función para agregar datos sin sobrescribir y preservar comentarios
        def append_to_excel(file_path, new_data, sheet_name, comment_text, is_spectro=False):
            try:
                # Cargar el libro de trabajo existente si existe
                if os.path.exists(file_path):
                    # Leer datos existentes
                    existing_df = pd.read_excel(file_path, sheet_name=sheet_name)
                    # Cargar el libro para leer los comentarios
                    book = openpyxl.load_workbook(file_path)
                    sheet = book[sheet_name]
                    # Guardar los comentarios existentes
                    existing_comments = {}
                    for row in sheet.rows:
                        for cell in row:
                            if cell.comment:
                                existing_comments[(cell.row, cell.column)] = cell.comment

                    # Verificar si ya existe la fila de longitudes de onda (solo para espectro)
                    has_wavelengths = False
                    if is_spectro:
                        # Comprobar si la segunda fila (índice 1 en pandas) tiene los valores de longitud de onda
                        if len(existing_df) > 0:
                            first_row = existing_df.iloc[0].tolist()
                            # Comparar los primeros valores para evitar problemas de precisión
                            if all(abs(first_row[i] - WAVELENGTHS[i]) < 0.01 for i in range(min(len(first_row), len(WAVELENGTHS)))):
                                has_wavelengths = True

                    # Agregar una fila separadora (fila vacía)
                    separator_row = pd.DataFrame([{}])
                    # Concatenar datos existentes, separador y nuevos datos
                    updated_df = pd.concat([existing_df, separator_row, new_data], ignore_index=True)
                    # Calcular la fila donde comienza el nuevo lote (después de la fila separadora)
                    first_row_of_new_batch = len(existing_df) + 2  # +1 por la fila separadora, +1 porque Excel comienza en 1
                else:
                    # Si el archivo no existe, usar solo los nuevos datos
                    updated_df = new_data
                    existing_comments = {}
                    has_wavelengths = False
                    first_row_of_new_batch = 2  # La primera fila de datos (después del encabezado)

                # Si es el archivo de espectro y no tiene las longitudes de onda, agregarlas al principio
                if is_spectro and not has_wavelengths:
                    wavelengths_df = pd.DataFrame([WAVELENGTHS])
                    updated_df = pd.concat([wavelengths_df, updated_df], ignore_index=True)
                    # Ajustar la fila donde comienza el nuevo lote
                    first_row_of_new_batch += 1  # Se agregó una fila más (las longitudes de onda)

                # Guardar el DataFrame actualizado
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    updated_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    # Acceder a la hoja de trabajo
                    worksheet = writer.sheets[sheet_name]

                    # Restaurar los comentarios existentes
                    for (row, col), comment in existing_comments.items():
                        worksheet.cell(row=row, column=col).comment = comment

                    # Agregar el comentario de las longitudes de onda (solo la primera vez)
                    if is_spectro and not has_wavelengths:
                        wavelength_comment = Comment("valores longitud de onda nm", "System")
                        worksheet.cell(row=2, column=1).comment = wavelength_comment  # Fila 2 porque la fila 1 es el encabezado

                    # Agregar el nuevo comentario en la primera fila del nuevo lote
                    comment = Comment(comment_text, "System")
                    worksheet.cell(row=first_row_of_new_batch, column=1).comment = comment

            except PermissionError:
                print(f"Error: No se puede acceder al archivo '{file_path}'. Asegúrate de que no esté abierto en otro programa.")

        # Guardar cada DataFrame en un archivo Excel y agregar el comentario
        append_to_excel(espectro_file, espectro_df, "Datos", f"=== {tipo} - {timestamp} ===", is_spectro=True)
        append_to_excel(temp_file, temp_df, "Datos", f"=== {tipo} - {timestamp} ===")
        append_to_excel(ph_file, ph_df, "Datos", f"=== {tipo} - {timestamp} ===")

        print(f"Lote de 10 mediciones guardado: {tipo} - {timestamp}")

    def save_manual_batch(self, timestamp):
        # Crear DataFrame para el espectrómetro
        espectro_data = []
        for espectrometro, _, _ in self.manual_batch:
            espectro_data.append(espectrometro)
        espectro_df = pd.DataFrame(espectro_data)

        # Crear DataFrame para la temperatura
        temp_data = [temp for _, temp, _ in self.manual_batch]
        temp_df = pd.DataFrame(temp_data, columns=["Temperatura"])

        # Crear DataFrame para el pH
        ph_data = [ph for _, _, ph in self.manual_batch]
        ph_df = pd.DataFrame(ph_data, columns=["pH"])

        # Guardar en un archivo Excel
        manual_file = os.path.join(self.save_dir, f"manual_{self.manual_name}.xlsx")

        # Función para agregar datos sin sobrescribir y preservar comentarios
        def append_to_excel_manual(file_path, espectro_df, temp_df, ph_df, sheet_name, comment_text):
            try:
                # Cargar el libro de trabajo existente si existe
                if os.path.exists(file_path):
                    # Leer datos existentes
                    existing_df = pd.read_excel(file_path, sheet_name=sheet_name)
                    # Cargar el libro para leer los comentarios
                    book = openpyxl.load_workbook(file_path)
                    sheet = book[sheet_name]
                    # Guardar los comentarios existentes
                    existing_comments = {}
                    for row in sheet.rows:
                        for cell in row:
                            if cell.comment:
                                existing_comments[(cell.row, cell.column)] = cell.comment

                    # Verificar si ya existe la fila de longitudes de onda
                    has_wavelengths = False
                    if len(existing_df) > 0:
                        first_row = existing_df.iloc[0].tolist()
                        # Comparar los primeros valores para evitar problemas de precisión
                        if all(abs(first_row[i] - WAVELENGTHS[i]) < 0.01 for i in range(min(len(first_row), len(WAVELENGTHS)))):
                            has_wavelengths = True

                    # Preparar los datos para concatenar
                    updated_df = existing_df
                    start_row = len(existing_df) + 2  # +1 por la fila separadora, +1 porque Excel comienza en 1
                else:
                    # Si el archivo no existe, empezar desde cero
                    updated_df = pd.DataFrame()
                    existing_comments = {}
                    has_wavelengths = False
                    start_row = 2  # La primera fila de datos (después del encabezado)

                # Si no tiene las longitudes de onda, agregarlas al principio
                if not has_wavelengths:
                    wavelengths_df = pd.DataFrame([WAVELENGTHS])
                    updated_df = pd.concat([wavelengths_df, updated_df], ignore_index=True)
                    start_row += 1  # Ajustar la fila de inicio

                # Agregar una fila separadora (fila vacía)
                separator_row = pd.DataFrame([{}])
                updated_df = pd.concat([updated_df, separator_row], ignore_index=True)

                # Agregar los datos del espectrómetro
                espectro_start_row = start_row
                updated_df = pd.concat([updated_df, espectro_df], ignore_index=True)

                # Ajustar temp_df para que tenga el mismo número de columnas que espectro_df y los datos estén en la primera columna
                num_columns = espectro_df.shape[1]  # Número de columnas del espectrómetro (288)
                temp_adjusted = pd.DataFrame(index=temp_df.index, columns=range(num_columns))
                temp_adjusted.iloc[:, 0] = temp_df["Temperatura"]  # Colocar los datos en la primera columna
                temp_adjusted.iloc[:, 1:] = None  # Rellenar el resto con None

                # Agregar una fila separadora y los datos de temperatura
                updated_df = pd.concat([updated_df, separator_row], ignore_index=True)
                temp_start_row = len(updated_df) + 2
                updated_df = pd.concat([updated_df, temp_adjusted], ignore_index=True)

                # Ajustar ph_df para que tenga el mismo número de columnas que espectro_df y los datos estén en la primera columna
                ph_adjusted = pd.DataFrame(index=ph_df.index, columns=range(num_columns))
                ph_adjusted.iloc[:, 0] = ph_df["pH"]  # Colocar los datos en la primera columna
                ph_adjusted.iloc[:, 1:] = None  # Rellenar el resto con None

                # Agregar una fila separadora y los datos de pH
                updated_df = pd.concat([updated_df, separator_row], ignore_index=True)
                ph_start_row = len(updated_df) + 2
                updated_df = pd.concat([updated_df, ph_adjusted], ignore_index=True)

                # Guardar el DataFrame actualizado
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    updated_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    # Acceder a la hoja de trabajo
                    worksheet = writer.sheets[sheet_name]

                    # Restaurar los comentarios existentes
                    for (row, col), comment in existing_comments.items():
                        worksheet.cell(row=row, column=col).comment = comment

                    # Agregar el comentario de las longitudes de onda (solo la primera vez)
                    if not has_wavelengths:
                        wavelength_comment = Comment("valores longitud de onda nm", "System")
                        worksheet.cell(row=2, column=1).comment = wavelength_comment  # Fila 2 porque la fila 1 es el encabezado

                    # Agregar comentarios para cada sección
                    espectro_comment = Comment(f"=== MUESTRA MANUAL - {timestamp} - Espectrómetro ===", "System")
                    worksheet.cell(row=espectro_start_row, column=1).comment = espectro_comment

                    temp_comment = Comment(f"=== MUESTRA MANUAL - {timestamp} - Temperatura ===", "System")
                    worksheet.cell(row=temp_start_row, column=1).comment = temp_comment

                    ph_comment = Comment(f"=== MUESTRA MANUAL - {timestamp} - pH ===", "System")
                    worksheet.cell(row=ph_start_row, column=1).comment = ph_comment

            except PermissionError:
                print(f"Error: No se puede acceder al archivo '{file_path}'. Asegúrate de que no esté abierto en otro programa.")

        # Guardar todos los datos en una sola hoja
        append_to_excel_manual(manual_file, espectro_df, temp_df, ph_df, "Datos", f"=== MUESTRA MANUAL - {timestamp} ===")

        print(f"Muestra manual de 10 mediciones guardada como: {self.manual_name}")

    def send_command(self, command):
        if self.serial_port and self.serial_port.is_open:
            if command.strip().upper() == 'M':
                self.waiting_for_manual = True
                self.manual_batch = []
                self.manual_name = input("Introduce un nombre para esta muestra manual: ").strip()
                print(f"Enviando 'M' al Arduino para '{self.manual_name}'...")
                self.serial_port.write("M\n".encode('utf-8'))
            else:
                self.serial_port.write(f"{command}\n".encode('utf-8'))

    def get_latest_measurements(self):
        """Return the latest batch of measurements for the dashboard."""
        return {
            "current_batch": self.current_batch,
            "manual_batch": self.manual_batch,
            "measurement_count": self.measurement_count,
            "is_first_measurement": self.is_first_measurement,
        }

def main():
    monitor = SerialMonitor(PORT, BAUD_RATE)
    monitor.set_save_directory()
    monitor.start()

    try:
        print(" Escribe comandos para enviar al Arduino (Ctrl+C para salir):")
        while True:
            comando = input()
            monitor.send_command(comando)
    except KeyboardInterrupt:
        print("\nInterrupción por usuario.")
    finally:
        monitor.stop()

if __name__ == "__main__":
    main()